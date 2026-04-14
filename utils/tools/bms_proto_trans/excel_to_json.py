#!/usr/bin/env python3
# -*- coding: utf-8 -*-

"""
Excel(单表: 每行一个Signal) -> JSON 协议导出器

你的表头（大小写/空格/换行不敏感）示例：
No, MsgName, MsgID (Hex), DLC, Vtype, Sender, Tx, Cycle time (ms),
Signal Name, Startbit（LSB）, Length, Factor, Offset, Initial Value,
Unit, Min, Max, Receiver, Comment, VALTable

特点：
- Excel 只有一张表（默认取第一个sheet，也可指定sheet）
- 每一行定义一个 signal
- 同一 MsgName 多行：聚合到同一个 message
- VALTable 是内联字符串，如： 255 "Signal Invalid"  或  0 "Off" 1 "On"
"""

import argparse
import json
import re
from datetime import datetime
from typing import Any, Dict, List, Optional

from openpyxl import load_workbook


# ---------- 基础转换工具 ----------
def _norm_col(s: Any) -> str:
    """把表头归一化：去空白/换行/括号/中文全角括号差异，统一小写"""
    if s is None:
        return ""
    s = str(s).strip().lower()
    # 把各种空白去掉
    s = re.sub(r"\s+", "", s)
    # 常见全角符号归一
    s = s.replace("（", "(").replace("）", ")").replace("：", ":")
    return s


def _to_str(x: Any) -> str:
    return "" if x is None else str(x).strip()


def _to_int(x: Any) -> Optional[int]:
    if x is None:
        return None
    if isinstance(x, int):
        return x
    if isinstance(x, float):
        return int(x)
    s = str(x).strip()
    if s == "":
        return None
    if s.lower().startswith("0x"):
        return int(s, 16)
    return int(float(s))


def _to_float(x: Any) -> Optional[float]:
    if x is None:
        return None
    if isinstance(x, (int, float)):
        return float(x)
    s = str(x).strip()
    if s == "":
        return None
    return float(s)


def _parse_byte_order(vtype: Any) -> str:
    s = _to_str(vtype).lower()
    if s == "":
        return "Intel"
    if "intel" in s or "little" in s or "lsb" in s:
        return "Intel"
    if "motorola" in s or "big" in s or "msb" in s:
        return "Motorola"
    # 容错：有些表可能写 0/1
    if s in ("0", "le"):
        return "Intel"
    if s in ("1", "be"):
        return "Motorola"
    # 兜底：原样首字母大写
    return "Intel" if s == "intel" else "Motorola" if s == "motorola" else "Intel"


# ---------- 读取单sheet为dict行 ----------
def _read_sheet_as_dicts(wb, sheet_name: str) -> List[Dict[str, Any]]:
    if sheet_name not in wb.sheetnames:
        raise ValueError(f"Sheet '{sheet_name}' not found. Existing: {wb.sheetnames}")
    ws = wb[sheet_name]
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return []

    header = rows[0]
    col_map: Dict[str, int] = {}
    for idx, h in enumerate(header):
        key = _norm_col(h)
        if key:
            col_map[key] = idx

    out: List[Dict[str, Any]] = []
    for r in rows[1:]:
        if r is None:
            continue
        if all(v is None or str(v).strip() == "" for v in r):
            continue
        d: Dict[str, Any] = {}
        for k, idx in col_map.items():
            d[k] = r[idx] if idx < len(r) else None
        out.append(d)
    return out


# ---------- VALTable 解析 ----------
_VAL_PAIR_RE = re.compile(r"""
    (?P<num>[-+]?\d+|0x[0-9a-fA-F]+)      # 数字：255 或 0xFF
    \s*                                  # 可选空白
    (?P<q>["“”'])                        # 引号开始（支持中文引号/单引号/双引号）
    (?P<txt>.*?)
    (?P=q)                               # 对称引号结束
""", re.VERBOSE)


def parse_val_table(cell: Any) -> Optional[Dict[str, str]]:
    """
    输入示例：
      255 "Signal Invalid"
      0 "Off" 1 "On" 2 "Fault"
      0xFF "Invalid"
    输出：
      {"255":"Signal Invalid"} / {"0":"Off","1":"On"} / {"255":"Invalid"}
    """
    s = _to_str(cell)
    if s == "":
        return None

    # 统一中文引号为双引号（便于解析）
    s = s.replace("“", '"').replace("”", '"')

    pairs = []
    for m in _VAL_PAIR_RE.finditer(s):
        num_s = m.group("num")
        txt = m.group("txt").strip()
        # 数字归一为十进制字符串 key（更利于后续 compare）
        if num_s.lower().startswith("0x"):
            key = str(int(num_s, 16))
        else:
            key = str(int(num_s))
        pairs.append((key, txt))

    if not pairs:
        # 没匹配到 "数字+引号文本" 的格式，给个保守处理：不报错但不生成表
        return None

    out: Dict[str, str] = {}
    for k, v in pairs:
        out[k] = v
    return out


# ---------- 主逻辑：单表聚合成 messages ----------
def excel_to_json_single_table(excel_path: str, sheet: Optional[str] = None) -> Dict[str, Any]:
    wb = load_workbook(excel_path, data_only=True)
    sheet_name = sheet or wb.sheetnames[0]
    rows = _read_sheet_as_dicts(wb, sheet_name)

    # 列名映射（归一化后）
    # 你表里是：MsgName, MsgID (Hex), DLC, Vtype, Sender, Tx, Cycle time (ms),
    # Signal Name, Startbit（LSB）, Length, Factor, Offset, Initial Value,
    # Unit, Min, Max, Receiver, Comment, VALTable
    COL = {
        "msg_name": "msgname",
        "msg_id": "msgid(hex)",
        "dlc": "dlc",
        "vtype": "vtype",
        "sender": "sender",
        "tx": "tx",
        "cycle": "cycletime(ms)",
        "sig_name": "signalname",
        "startbit": "startbit(lsb)",          # 注意：你的截图是 Startbit（LSB）
        "length": "length",
        "factor": "factor",
        "offset": "offset",
        "init": "initialvalue",
        "unit": "unit",
        "min": "min",
        "max": "max",
        "receiver": "receiver",
        "comment": "comment",
        "valtable": "valtable",
    }

    msg_map: Dict[str, Dict[str, Any]] = {}

    for i, r in enumerate(rows):
        msg_name = _to_str(r.get(COL["msg_name"]))
        if msg_name == "":
            raise ValueError(f"Row {i+2}: MsgName empty")

        msg_id = _to_int(r.get(COL["msg_id"]))
        if msg_id is None:
            raise ValueError(f"Row {i+2}: MsgID (Hex) empty for MsgName={msg_name}")

        dlc = _to_int(r.get(COL["dlc"])) or 8
        bo = _parse_byte_order(r.get(COL["vtype"]))
        cycle = _to_int(r.get(COL["cycle"]))
        sender = _to_str(r.get(COL["sender"]))
        receiver = _to_str(r.get(COL["receiver"]))
        tx_type = _to_str(r.get(COL["tx"]))
        comment_msg = ""  # 这里没有独立 message comment，先空着

        m = msg_map.get(msg_name)
        if m is None:
            m = {
                "name": msg_name,
                "id": int(msg_id),
                "id_hex": f"0x{int(msg_id):08X}",
                "dlc": int(dlc),
                "byte_order": bo,
                "cycle_ms": int(cycle) if cycle is not None else None,
                "sender": sender,
                "receiver": receiver,
                "tx": tx_type,
                "comment": comment_msg,
                "signals": [],
            }
            msg_map[msg_name] = m
        else:
            # 一致性检查（避免表里同名消息写错）
            if int(msg_id) != m["id"]:
                raise ValueError(f"Row {i+2}: MsgID mismatch for {msg_name}: {m['id_hex']} vs 0x{int(msg_id):08X}")
            if int(dlc) != m["dlc"]:
                raise ValueError(f"Row {i+2}: DLC mismatch for {msg_name}")
            if bo != m["byte_order"]:
                raise ValueError(f"Row {i+2}: Vtype/byte_order mismatch for {msg_name}")

        # signal
        sig_name = _to_str(r.get(COL["sig_name"]))
        if sig_name == "":
            raise ValueError(f"Row {i+2}: Signal Name empty for MsgName={msg_name}")

        startbit = _to_int(r.get(COL["startbit"]))
        length = _to_int(r.get(COL["length"]))
        if startbit is None or length is None:
            raise ValueError(f"Row {i+2}: Startbit/Length missing for {msg_name}.{sig_name}")

        factor = _to_float(r.get(COL["factor"]))
        offset = _to_float(r.get(COL["offset"]))

        sig_def: Dict[str, Any] = {
            "name": sig_name,
            "startbit_lsb": int(startbit),
            "length": int(length),
            "factor": float(factor) if factor is not None else 1.0,
            "offset": float(offset) if offset is not None else 0.0,
            "initial": _to_float(r.get(COL["init"])),
            "unit": _to_str(r.get(COL["unit"])),
            "min": _to_float(r.get(COL["min"])),
            "max": _to_float(r.get(COL["max"])),
            "receiver": receiver,
            "comment": _to_str(r.get(COL["comment"])),
        }

        vt = parse_val_table(r.get(COL["valtable"]))
        if vt:
            sig_def["val_table"] = vt

        m["signals"].append(sig_def)

    out_msgs = list(msg_map.values())
    # 让输出稳定：按 id 升序
    out_msgs.sort(key=lambda x: x["id"])

    return {
        "meta": {
            "source": excel_path,
            "sheet": sheet_name,
            "generated_at": datetime.utcnow().isoformat() + "Z",
        },
        "messages": out_msgs,
    }


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("excel", help="input xlsx path")
    ap.add_argument("output", help="output json path")
    ap.add_argument("--sheet", default=None, help="sheet name (default: first sheet)")
    ap.add_argument("--indent", type=int, default=2)
    args = ap.parse_args()

    proto = excel_to_json_single_table(args.excel, args.sheet)
    with open(args.output, "w", encoding="utf-8") as f:
        json.dump(proto, f, ensure_ascii=False, indent=args.indent)

    print(f"[OK] Wrote: {args.output}")
    print(f"     messages={len(proto.get('messages', []))}")


if __name__ == "__main__":
    main()

# python3 utils/tools/excel_to_json.py config/protocol/msg.xlsx config/protocol/bms_j1939.json